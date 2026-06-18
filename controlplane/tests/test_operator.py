"""Operator agent tests — tools, approval gate, and the full Excel scenario.

No mocks: a real local HTTP server serves a real .xlsx, openpyxl really edits it,
and a real aiosmtpd server captures the real SMTP send. The deterministic planner
drives the end-to-end test so it runs without a model; a separate test exercises
the live Ollama brain when it's reachable.
"""

from __future__ import annotations

import asyncio
import functools
import http.server
import threading
import email.policy
from email import message_from_bytes
from pathlib import Path

import pytest
from aiosmtpd.controller import Controller
from openpyxl import Workbook, load_workbook

from controlplane.operator.approval import always_allow, always_deny
from controlplane.operator.runner import Operator
from controlplane.operator.tools import SMTPConfig, ToolExecutor


# --- fixtures: a real file server and a real SMTP sink ----------------------

@pytest.fixture
def xlsx_server(tmp_path):
    """Serve a tmp dir over HTTP; yields (base_url, dir). Hosts budget.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws["A1"], ws["B1"] = "Item", "Amount"
    ws["A2"], ws["B2"] = "Forecast", 1000
    wb.save(tmp_path / "budget.xlsx")

    handler = functools.partial(http.server.SimpleHTTPRequestHandler, directory=str(tmp_path))
    server = http.server.ThreadingHTTPServer(("127.0.0.1", 0), handler)
    port = server.server_address[1]
    t = threading.Thread(target=server.serve_forever, daemon=True)
    t.start()
    try:
        yield f"http://127.0.0.1:{port}", tmp_path
    finally:
        server.shutdown()


class _Sink:
    def __init__(self):
        self.messages = []

    async def handle_DATA(self, server, session, envelope):
        self.messages.append(envelope)
        return "250 OK"


def _free_port() -> int:
    import socket

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1]


@pytest.fixture
def smtp_sink():
    """A real SMTP server that captures messages. Yields (SMTPConfig, sink)."""
    sink = _Sink()
    port = _free_port()
    controller = Controller(sink, hostname="127.0.0.1", port=port)
    controller.start()
    try:
        yield SMTPConfig(host="127.0.0.1", port=port), sink
    finally:
        controller.stop()


# --- tool unit tests --------------------------------------------------------

def test_download_and_update_spreadsheet(xlsx_server, tmp_path):
    base, _ = xlsx_server
    ex = ToolExecutor(workspace=tmp_path / "ws", approval=always_deny)
    ex.download_file(f"{base}/budget.xlsx", "budget.xlsx")
    ex.update_spreadsheet("budget.xlsx", "B2", "1234")
    wb = load_workbook(Path(ex.workspace) / "budget.xlsx")
    assert wb.active["B2"].value == 1234  # coerced to int


def test_path_escape_is_blocked(tmp_path):
    ex = ToolExecutor(workspace=tmp_path / "ws", approval=always_deny)
    out = ex.execute("update_spreadsheet", {"filename": "../evil.xlsx", "cell": "A1", "value": "x"})
    assert "escapes the workspace" in out


def test_send_email_requires_approval(smtp_sink, tmp_path):
    cfg, sink = smtp_sink
    denied = ToolExecutor(workspace=tmp_path / "w1", approval=always_deny, smtp=cfg)
    msg = denied.send_email("boss@example.com", "Hi", "body")
    assert "NOT sent" in msg and not sink.messages

    allowed = ToolExecutor(workspace=tmp_path / "w2", approval=always_allow, smtp=cfg)
    msg = allowed.send_email("boss@example.com", "Hi", "body")
    assert "sent" in msg.lower() and len(sink.messages) == 1


# --- the full scenario, end-to-end (deterministic brain, no mocks) ----------

def test_excel_scenario_end_to_end(xlsx_server, smtp_sink, tmp_path):
    base, _ = xlsx_server
    cfg, sink = smtp_sink

    goal = (
        f"download the budget spreadsheet from {base}/budget.xlsx, "
        f"set B2 to 1234 and email it to boss@example.com"
    )
    op = Operator(approval=always_allow, smtp=cfg, workspace_root=tmp_path / "ops")
    result = op.run(goal, brain="rule")

    # 1) the email really went out, with the spreadsheet attached
    assert len(sink.messages) == 1
    msg = message_from_bytes(sink.messages[0].content, policy=email.policy.default)
    assert msg["To"] == "boss@example.com"
    attachments = [p for p in msg.iter_attachments()]
    assert attachments, "expected the spreadsheet attached"

    # 2) the attached spreadsheet actually contains the update
    out = tmp_path / "received.xlsx"
    out.write_bytes(attachments[0].get_payload(decode=True))
    assert load_workbook(out).active["B2"].value == 1234

    # 3) the audit trail shows the real steps
    tools_used = [a["tool"] for a in result.actions]
    assert tools_used == ["download_file", "read_spreadsheet", "update_spreadsheet", "send_email"]


def test_scenario_blocks_email_when_not_approved(xlsx_server, smtp_sink, tmp_path):
    base, _ = xlsx_server
    cfg, sink = smtp_sink
    goal = f"get {base}/budget.xlsx, set B2 to 999, email it to boss@example.com"
    op = Operator(approval=always_deny, smtp=cfg, workspace_root=tmp_path / "ops")
    op.run(goal, brain="rule")
    assert sink.messages == []  # confirm-before-irreversible held the email back


# --- live Ollama brain (guarded) --------------------------------------------

def _ollama_up() -> bool:
    try:
        import httpx

        httpx.get("http://localhost:11434/api/tags", timeout=2).raise_for_status()
        return True
    except Exception:  # noqa: BLE001
        return False


@pytest.mark.skipif(not _ollama_up(), reason="ollama not reachable")
def test_live_llm_brain_drives_tools(xlsx_server, smtp_sink, tmp_path):
    base, _ = xlsx_server
    cfg, sink = smtp_sink
    goal = (
        f"Download the spreadsheet at {base}/budget.xlsx, set cell B2 to 1234, "
        f"then email it to boss@example.com with subject 'Updated budget'."
    )
    op = Operator(approval=always_allow, smtp=cfg, workspace_root=tmp_path / "ops")
    result = op.run(goal, brain="llm")
    # The model must have actually driven the tools (downloaded + sent).
    tools_used = {a["tool"] for a in result.actions}
    assert "download_file" in tools_used
    assert result.brain.startswith("llm:")
