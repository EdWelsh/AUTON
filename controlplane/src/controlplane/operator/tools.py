"""Real, headless tools the operator agent calls to get work done.

Every tool does a real thing — downloads a file, edits a spreadsheet, sends an
email — confined to a per-task workspace directory. Irreversible tools (email)
pass through an approval gate before acting. Tool schemas are OpenAI/LiteLLM
function-calling shaped so any provider (Ollama, OpenAI, Anthropic) can drive them.
"""

from __future__ import annotations

import json
import smtplib
from dataclasses import dataclass
from email.message import EmailMessage
from pathlib import Path
from typing import Callable

import httpx
from openpyxl import load_workbook

# An approval gate: (action, detail) -> approved?  Irreversible tools call it.
Approval = Callable[[str, str], bool]


@dataclass(frozen=True)
class SMTPConfig:
    host: str = "localhost"
    port: int = 1025
    sender: str = "auton@localhost"
    username: str | None = None
    password: str | None = None
    use_tls: bool = False


class ToolError(Exception):
    """Raised inside a tool; the runner turns it into an observation for the model."""


class ToolExecutor:
    """Executes named tool calls against a sandboxed workspace."""

    def __init__(
        self,
        workspace: Path,
        approval: Approval,
        smtp: SMTPConfig | None = None,
    ) -> None:
        self.workspace = Path(workspace)
        self.workspace.mkdir(parents=True, exist_ok=True)
        self.approval = approval
        self.smtp = smtp or SMTPConfig()
        # Audit trail of everything the agent did (for the transcript / tests).
        self.actions: list[dict] = []

    # --- path safety --------------------------------------------------------

    def _resolve(self, filename: str) -> Path:
        """Resolve a filename strictly inside the workspace (no escaping)."""
        p = (self.workspace / filename).resolve()
        if self.workspace.resolve() not in p.parents and p != self.workspace.resolve():
            raise ToolError(f"path {filename!r} escapes the workspace")
        return p

    # --- tools --------------------------------------------------------------

    def download_file(self, url: str, filename: str) -> str:
        dest = self._resolve(filename)
        with httpx.Client(follow_redirects=True, timeout=30) as client:
            r = client.get(url)
            r.raise_for_status()
            dest.write_bytes(r.content)
        self.actions.append({"tool": "download_file", "url": url, "saved": str(dest), "bytes": len(r.content)})
        return f"Downloaded {len(r.content)} bytes from {url} to {filename}"

    def read_spreadsheet(self, filename: str, max_rows: int = 20) -> str:
        path = self._resolve(filename)
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([("" if c is None else c) for c in row])
        self.actions.append({"tool": "read_spreadsheet", "file": filename, "sheet": ws.title})
        return f"Sheet '{ws.title}' ({ws.max_row}x{ws.max_column}):\n" + json.dumps(rows)

    def update_spreadsheet(self, filename: str, cell: str, value: str) -> str:
        path = self._resolve(filename)
        wb = load_workbook(path)
        ws = wb.active
        # Coerce numeric strings to numbers so totals/formulas behave.
        coerced: object = value
        try:
            coerced = int(value)
        except (TypeError, ValueError):
            try:
                coerced = float(value)
            except (TypeError, ValueError):
                coerced = value
        ws[cell] = coerced
        wb.save(path)
        self.actions.append({"tool": "update_spreadsheet", "file": filename, "cell": cell, "value": coerced})
        return f"Set {cell} = {coerced!r} in {filename} (sheet '{ws.title}')"

    def send_email(self, to: str, subject: str, body: str, attachment: str | None = None) -> str:
        # Irreversible — gate on approval before touching the network.
        detail = f"To: {to}\nSubject: {subject}\nAttachment: {attachment or 'none'}\n\n{body}"
        if not self.approval("send_email", detail):
            self.actions.append({"tool": "send_email", "to": to, "status": "declined"})
            return "Email NOT sent — you declined the confirmation."

        msg = EmailMessage()
        msg["From"] = self.smtp.sender
        msg["To"] = to
        msg["Subject"] = subject
        msg.set_content(body)
        if attachment:
            path = self._resolve(attachment)
            data = path.read_bytes()
            msg.add_attachment(
                data,
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=path.name,
            )
        with smtplib.SMTP(self.smtp.host, self.smtp.port, timeout=20) as server:
            if self.smtp.use_tls:
                server.starttls()
            if self.smtp.username:
                server.login(self.smtp.username, self.smtp.password or "")
            server.send_message(msg)
        self.actions.append({"tool": "send_email", "to": to, "subject": subject, "status": "sent"})
        return f"Email sent to {to} (subject: {subject})"

    # --- dispatch -----------------------------------------------------------

    def execute(self, name: str, args: dict) -> str:
        """Run a tool by name, returning an observation string (errors included)."""
        fn = {
            "download_file": self.download_file,
            "read_spreadsheet": self.read_spreadsheet,
            "update_spreadsheet": self.update_spreadsheet,
            "send_email": self.send_email,
        }.get(name)
        if fn is None:
            return f"ERROR: unknown tool {name!r}"
        try:
            return fn(**args)
        except Exception as exc:  # noqa: BLE001 - surface as an observation, never crash the loop
            return f"ERROR running {name}: {exc}"


def tool_schemas() -> list[dict]:
    """OpenAI/LiteLLM function schemas for the operator's tools."""
    return [
        {
            "type": "function",
            "function": {
                "name": "download_file",
                "description": "Download a file from a URL into the task workspace.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string"},
                        "filename": {"type": "string", "description": "name to save as in the workspace"},
                    },
                    "required": ["url", "filename"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "read_spreadsheet",
                "description": "Read the active sheet of an .xlsx file to inspect its contents.",
                "parameters": {
                    "type": "object",
                    "properties": {"filename": {"type": "string"}},
                    "required": ["filename"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "update_spreadsheet",
                "description": "Set a single cell (e.g. 'B2') to a value in an .xlsx file.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "filename": {"type": "string"},
                        "cell": {"type": "string", "description": "A1-style cell reference, e.g. B2"},
                        "value": {"type": "string"},
                    },
                    "required": ["filename", "cell", "value"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "send_email",
                "description": "Send an email (REQUIRES user confirmation). Optionally attach a workspace file.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "to": {"type": "string"},
                        "subject": {"type": "string"},
                        "body": {"type": "string"},
                        "attachment": {"type": "string", "description": "workspace filename to attach"},
                    },
                    "required": ["to", "subject", "body"],
                },
            },
        },
    ]
